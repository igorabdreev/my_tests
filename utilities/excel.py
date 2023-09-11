from uuid import uuid4

from allure import attach, step
from allure_commons.types import AttachmentType
from openpyxl import Workbook, load_workbook

from utilities.config import Config
from utilities.logging import logger


class Xlsx:

    def __init__(self, filename: str):
        self.filename: str = filename
        self._book: Workbook = self._create_excel_obj()
        self.path_folder = Config.xlsx_dir / str(uuid4())
        self.path: str = str(self.path_folder / self.filename)
        logger.debug(
            f'\tНазвание файла: {self.filename}\n'
            f'\tXlsx книга:     {self._book}\n'
            f'\tПуть до папки:  {self.path_folder}\n'
            f'\tПуть до файла:  {self.path}'
        )

    @step(f'Создать книгу Excel')
    def _create_excel_obj(self) -> Workbook:
        logger.info(f'Создание книги Excel')
        return Workbook()

    @step('Сохранить Excel файл')
    def _save_xlsx(self) -> str:
        """Сохранить книгу в файл по указанному пути"""
        if not self.path_folder.exists():
            with step(msg := f'Создать путь до папки с xlsx "{self.path_folder}"'):
                self.path_folder.mkdir(parents=True)
                logger.info(msg)

        with step(msg := f'Сохранить книгу в файл по пути {self.path}'):
            self._book.save(filename=self.path)
            logger.info(msg)

        return self.path.replace(str(Config.test_data_dir), '')

    @staticmethod
    def open(filename: str) -> Workbook:
        """Открыть файл xlsx, распарсить и получить объект книги

        Args:
            filename: путь до файла
        """
        with step(msg := f'Открыть файл xlsx "{(path := Config.test_data_dir / filename)}"'):
            logger.info(msg)
            return load_workbook(filename=path)

    @step('Добавить записи в лист по умолчанию в книгу Excel')
    def add_to_book(self, values: dict):
        """ Добавить записи в лист по умолчанию в книгу Excel

        Args:
            values: словарь с адресами ячеек (ключами) и их значеними
        """
        logger.info(f'Запись значений в активный лист книги Excel')
        message: str = ''

        for key, value in values.items():
            self._book.active[key] = value
            message += (msg := f'Значение "{value}" записано в ячейку "{key}"') + '\n'
            logger.debug(msg)

        attach(body=message, name='Запись в EXCEL', attachment_type=AttachmentType.TEXT)

    def __enter__(self):
        """Контекстный менеджер для работы с XLSX файлом"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Контекстный менеджер для прекращения работы с XLSX файлом и его сохранения"""
        self._save_xlsx()
